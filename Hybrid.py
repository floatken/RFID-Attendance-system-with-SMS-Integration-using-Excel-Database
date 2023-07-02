#Import Libraries
import openpyxl
import pandas as pd
from openpyxl.styles import PatternFill
from datetime import datetime, timedelta
import os
import serial
import time

#Wait for the Orangepi/Raspberrypi to correct time
#time.sleep(300)

#Get the current date
totoo = True
now = None
date_string = None
while totoo == True:
    
    now = datetime.now()
    date_string = now.strftime("%Y-%m-%d")
    totoo = False

#SMS Test Function
def send_TEST(phone_number, message):
 serial_port = 'COM3'  
 baud_rate = 57600  
 ser = serial.Serial(serial_port, baud_rate, timeout=1)

# Send AT command to check if the module is responding
 ser.write(b'AT\r\n')
 response = ser.read(100).decode()
 print(response)

 # Set the SMS message format to text mode
 ser.write(b'AT+CMGF=1\r\n')
 response = ser.read(100).decode()
 print(response)

 # Set the recipient's phone number
 # Send the SMS
 ser.write('AT+CMGS="{}"\r\n'.format(phone_number).encode())
 response = ser.read(100).decode()
 print(response)
 ser.write(message.encode() + b'\r\n')
 ser.write(bytes([26]))  # ASCII code for Ctrl+Z
 response = ser.read(100).decode()
 print(response)

 # Close the serial connection
 ser.close()
phone_number = '639202734994' #Input working phone number
message = f'Test {now}' 
send_TEST(phone_number, message)

#SMS Function
def send_message(phone_number, message):
    ser = serial.Serial('COM3', 57600, timeout=1)
    ser.write(b'AT\r\n')
    response = ser.read(100)

    if b'OK' not in response:
        # Module is not ready, handle the error or raise an exception
        return False

    # Construct the message command
    message_cmd = f'AT+CMGS="{phone_number}"\r\n'

    attempt = 1
    max_attempts = 3

    while attempt <= max_attempts:
        # Send the message command
        ser.write(message_cmd.encode())
        response = ser.read(100)

        if b'>' in response:
            # Send the message content
            ser.write(message.encode())
            ser.write(b'\x1A')

            # Wait for the response
            response = ser.read(100)

            if b'+CMGS:' in response:
                # Message sent successfully
                return True
               
        # Increment the attempt counter
        attempt += 1

    # Message sending failed after max_attempts
    
    return False

# Function to clear recent scanned RFIDs
def clear_rfids(): 
    simula = 1
    tapos = ws.max_row
    for row in ws.iter_rows(min_row=simula, max_row=tapos, min_col=1, max_col=2):
        for cell in row:
            cell.value = None
    wb.save(filename)

# Funtion to record RFID
def RFID_recording_in(id, timestamp):
    
    # add data to attendance list
    ws.append([id, timestamp])
    wm = wb["Sheet"]
    last = wm.max_row
        
    # Set the number of preceding rows to check for duplicate RFID IDs
    preceding_rows = 10
    
    # Get the range of rows to iterate over
    start_row = max(last - preceding_rows, 1)  # Ensure not to go below the first row
    
    for row in ws.iter_rows(min_row=last, min_col=1, max_col=2, values_only=True):
        rfid_id = row[0]
        timez = row[1]
    
        # Check if the RFID ID has occurred in the preceding rows
        duplicate = False
        for i in range(start_row, last):
            if ws.cell(row=i, column=1).value == rfid_id:
                duplicate = True
                break
    
        if duplicate:
            continue  # Skip the iteration if duplicate RFID ID is found
        
        rfid_in = rfid_id
            
            # find the corresponding student in the Attendance Database
        student = database.loc[database["RFID ID"] == rfid_in]
        if len(student) > 0:
            name = student.iloc[0]["Name"]
            section = student.iloc[0]["Section"]
            phone_number = student.iloc[0]["Number"]
            message = f"This is an Automatic RFID Attendance System SMS message. {name} of {section} arrived in Pacita Complex Senior Highschool at {timez}"
            wp = wb[section]
            wp.append([name, timez])

                #Check if the student is late
            if datetime.strptime(timez, '%H:%M:%S').time() > datetime.strptime('07:15:00', '%H:%M:%S').time(): #set late time
                cell = wp.cell(wp.max_row, column=2)
                cell.fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
                message = f"This is an Automatic RFID Attendance System SMS message. {name} of {section} arrived late in Pacita Complex Senior Highschool at {timez}"
                send_message(phone_number, message)
                    
            send_message(phone_number, message)
                
        wb.save(filename)
        time.sleep(2)
def RFID_recording_out(id, timestamp):    
        # add data to attendance list
    ws.append([id, timestamp])
    wm = wb["Sheet"]
    last = wm.max_row
            
        # Set the number of preceding rows to check for duplicate RFID IDs
    preceding_rows = 10
        
        # Get the range of rows to iterate over
    start_row = max(last - preceding_rows, 1)  # Ensure not to go below the first row
        
    for row in ws.iter_rows(min_row=last, min_col=1, max_col=2, values_only=True):
        rfid_id = row[0]
        timez = row[1]
        
            # Check if the RFID ID has occurred in the preceding rows
        duplicate = False
        for i in range(start_row, last):
            if ws.cell(row=i, column=1).value == rfid_id:
                duplicate = True
                break

        if duplicate:
            continue  # Skip the iteration if duplicate RFID ID is found
        
        rfid_out = rfid_id
        timez = datetime.now().strftime("%H:%M:%S")
            
        student = database.loc[database["RFID ID"] == rfid_out]
        if len(student) > 0:
            name = student.iloc[0]["Name"]
            section = student.iloc[0]["Section"]
            phone_number = student.iloc[0]["Number"]
            message = f"This is an Automatic RFID Attendance System SMS. {name} of {section} left Pacita Complex Senior Highschool at {timez}"
                # write the attendance data to the corresponding sheet
            wp = wb[section]
            row_number = None
            for row_idx, row in enumerate(wp.iter_rows(values_only=True), start=1):
                if row[0] == name:
                    row_number = row_idx  # Get the row number
                    break

            # Check if the name was found in the sheet
            if row_number is None:
                print(f"{name}'s name was not found in the Excel file.")
            else:
                # Get the current timestamp
                

                # Update the "Time out" for the specified name
                wp.cell(row=row_number, column=3).value = timez
                send_message(phone_number, message)
                
                    
            
                # Save the modified Excel file
            wb.save(filename)
            time.sleep(2)

# Create or load the Attendance List file
filename = f"Attendance List {date_string}.xlsx"
if not os.path.isfile(filename):
    wv = openpyxl.Workbook()
    wv.save(filename)
global wb
global ws
wb = openpyxl.load_workbook(filename)
ws = wb.active

#Late indication for cells
red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')

#Load the Attendance Database into a pandas dataframe
database = pd.read_excel("Attendance Database.xlsx")

# get unique section names
sections = database["Section"].unique()

# Loop through each section and create a new sheet if it doesn't already exist
for section in sections:
    if section not in wb.sheetnames:
        wx = wb.create_sheet(section)
        wx.append(["Name", "Time In", "Time Out"])
database['RFID ID'] = "000" + database['RFID ID'].astype(str) #Code ignores the zeros in front of the RFID ID in the excel database so you need to add 3 zeros
# save the Attendance List file
wb.save(filename)

clear_rfids()

wb.save(filename)
arrival = True
sacrifice = None
while arrival == True:

    # get RFID id and timestamp
    timestamp = datetime.now().strftime("%H:%M:%S")

    if datetime.strptime(timestamp, '%H:%M:%S').time() < datetime.strptime('01:49:00', '%H:%M:%S').time(): #Set Dismissal time
        
        id = input("RFID:")
        sacrifice = id
        RFID_recording_in(id, timestamp)
            

    else:
        print("It's uwian time")
        arrival = False
        
#For the 1st scan of Dismissal
clear_rfids()
id = sacrifice
timestamp = datetime.now().strftime("%H:%M:%S")
RFID_recording_out(id, timestamp)


while True:

    # get RFID id and timestamp
    id = input("RFID:")
    timestamp = datetime.now().strftime("%H:%M:%S")
    RFID_recording_out(id, timestamp)